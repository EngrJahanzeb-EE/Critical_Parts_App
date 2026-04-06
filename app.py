import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Critical Parts Logger",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Global ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

/* ── Hide default Streamlit chrome ── */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: #0f1923;
    border-right: 1px solid #1e2d3d;
}
[data-testid="stSidebar"] * { color: #c9d8e8 !important; }
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stTextInput label { color: #7a9bb5 !important; font-size: 0.78rem; text-transform: uppercase; letter-spacing: 0.06em; }
[data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 { color: #e8f4fd !important; }

/* ── Sidebar inputs dark ── */
[data-testid="stSidebar"] input,
[data-testid="stSidebar"] .stSelectbox > div > div {
    background: #1a2635 !important;
    border-color: #2d4054 !important;
    color: #e8f4fd !important;
}

/* ── Main area card-style sections ── */
.part-card {
    background: #ffffff;
    border: 1px solid #e8eef4;
    border-radius: 10px;
    padding: 1.1rem 1.3rem;
    margin-bottom: 0.7rem;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
    transition: box-shadow 0.2s;
}
.part-card:hover { box-shadow: 0 3px 12px rgba(24,95,165,0.1); }

/* ── Type badges ── */
.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 4px;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 0.04em;
    text-transform: uppercase;
}
.badge-drive    { background:#e6f1fb; color:#0c447c; }
.badge-plc      { background:#eeedfe; color:#3c3489; }
.badge-encoder  { background:#eaf3de; color:#27500a; }
.badge-motor    { background:#faece7; color:#712b13; }
.badge-loadcell { background:#faeeda; color:#633806; }
.badge-hmi      { background:#fbeaf0; color:#72243e; }
.badge-relay    { background:#e1f5ee; color:#085041; }
.badge-breaker  { background:#d3d1c7; color:#2c2c2a; }
.badge-other    { background:#f1efe8; color:#444441; }

/* ── Metric boxes ── */
.metric-row {
    display: flex; gap: 12px; margin-bottom: 1.2rem;
}
.metric-box {
    flex: 1;
    background: #f4f8ff;
    border: 1px solid #d0e3f8;
    border-radius: 8px;
    padding: 0.9rem 1rem;
    text-align: center;
}
.metric-box .num { font-size: 1.8rem; font-weight: 600; color: #185fa5; }
.metric-box .lbl { font-size: 0.75rem; color: #6b8ba4; text-transform: uppercase; letter-spacing: 0.05em; }

/* ── Section headers ── */
.sec-header {
    font-size: 0.72rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: #8fa5bc;
    border-bottom: 1px solid #e4ecf3;
    padding-bottom: 6px;
    margin: 1.2rem 0 0.8rem;
}

/* ── Tag chips ── */
.tag-chip {
    display: inline-block;
    background: #eef3f8;
    border: 1px solid #d0dcea;
    border-radius: 4px;
    padding: 2px 8px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.75rem;
    color: #2a4a68;
    margin-right: 4px;
}

/* ── Empty state ── */
.empty-state {
    text-align: center;
    padding: 3rem 1rem;
    color: #b0c4d8;
}
.empty-state .icon { font-size: 2.5rem; margin-bottom: 0.5rem; }
.empty-state p { font-size: 0.9rem; }

/* ── Export button ── */
div[data-testid="stDownloadButton"] button {
    background: #185fa5 !important;
    color: white !important;
    border: none !important;
    width: 100%;
    padding: 0.65rem 1rem !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
    border-radius: 8px !important;
    letter-spacing: 0.02em;
}
div[data-testid="stDownloadButton"] button:hover {
    background: #0c447c !important;
}

/* ── Stacked form inputs tighter ── */
.stTextInput input, .stSelectbox select, .stNumberInput input {
    border-radius: 6px !important;
    font-size: 0.88rem !important;
}

/* ── Delete button red ── */
.del-btn button {
    background: #fff5f5 !important;
    border: 1px solid #ffc0c0 !important;
    color: #a32d2d !important;
    border-radius: 6px !important;
    font-size: 0.75rem !important;
    padding: 0.2rem 0.6rem !important;
}

/* ── Add button accent ── */
div[data-testid="stButton"] button[kind="primary"] {
    background: #185fa5 !important;
    color: white !important;
    border: none !important;
    border-radius: 7px !important;
    font-weight: 600 !important;
}

hr { border-color: #e8eef4; }
</style>
""", unsafe_allow_html=True)

# ── Session state init ────────────────────────────────────────────────────────
if "parts" not in st.session_state:
    st.session_state.parts = []   # list of dicts

# ── Component type definitions ────────────────────────────────────────────────
COMP_TYPES = [
    "VFD / Drive",
    "Motor",
    "PLC",
    "HMI",
    "Encoder / Resolver",
    "Load Cell",
    "Circuit Breaker / MCCB",
    "Relay / Contactor",
    "Transformer",
    "Other",
]

BADGE_CLASS = {
    "VFD / Drive":             "badge-drive",
    "Motor":                   "badge-motor",
    "PLC":                     "badge-plc",
    "HMI":                     "badge-hmi",
    "Encoder / Resolver":      "badge-encoder",
    "Load Cell":               "badge-loadcell",
    "Circuit Breaker / MCCB":  "badge-breaker",
    "Relay / Contactor":       "badge-relay",
    "Transformer":             "badge-other",
    "Other":                   "badge-other",
}

# ── Sidebar — Add Component ───────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚡ Critical Parts Logger")
    st.markdown("<div style='color:#4a7fa5;font-size:0.78rem;margin-bottom:1.2rem;'>Sapphire Fibres Limited</div>", unsafe_allow_html=True)
    st.divider()

    st.markdown("### Add New Component")

    department = st.text_input("Department", placeholder="e.g. Weaving, Spinning, Sizing")
    machine    = st.text_input("Machine / Location", placeholder="e.g. Loom-12, Warper-3")
    comp_type  = st.selectbox("Component Type", COMP_TYPES)

    st.divider()

    # ── Dynamic fields by type ────────────────────────────────────────────────
    tag_panel     = st.text_input("Tag on Panel", placeholder="e.g. VFD-01")
    tag_schematic = st.text_input("Tag on Schematic / SLD", placeholder="e.g. M1, K3")
    make          = st.text_input("Make / Brand", placeholder="e.g. Siemens, ABB, Schneider")
    model         = st.text_input("Model No.", placeholder="e.g. G120, ACS880")

    # Common fields shown per type
    kw = amps = voltage = rpm = None
    freq = ip_rating = ins_class = None
    capacity = poles = breaking_cap = None
    kva = prim_v = sec_v = None
    io_count = comm = ppr = output_type = None
    custom_rating = None

    if comp_type == "VFD / Drive":
        kw      = st.text_input("Power Rating (kW)", placeholder="e.g. 7.5")
        amps    = st.text_input("Output Current (A)", placeholder="e.g. 18.5")
        voltage = st.text_input("Input Voltage (V)", placeholder="e.g. 400")
        freq    = st.text_input("Frequency (Hz)", placeholder="e.g. 50")

    elif comp_type == "Motor":
        kw        = st.text_input("Power (kW / HP)", placeholder="e.g. 5.5 kW")
        amps      = st.text_input("Full Load Current (A)", placeholder="e.g. 12.5")
        voltage   = st.text_input("Voltage (V)", placeholder="e.g. 400")
        rpm       = st.text_input("Speed (RPM)", placeholder="e.g. 1450")
        poles     = st.text_input("No. of Poles", placeholder="e.g. 4")
        ins_class = st.text_input("Insulation Class", placeholder="e.g. F")
        ip_rating = st.text_input("IP Rating", placeholder="e.g. IP55")

    elif comp_type == "PLC":
        voltage  = st.text_input("Supply Voltage (V)", placeholder="e.g. 24 DC")
        io_count = st.text_input("I/O Count", placeholder="e.g. 32 DI / 16 DO")
        comm     = st.text_input("Comm. Protocol", placeholder="e.g. Profibus, Profinet")

    elif comp_type == "HMI":
        voltage = st.text_input("Supply Voltage (V)", placeholder="e.g. 24 DC")
        comm    = st.text_input("Comm. Protocol", placeholder="e.g. Profinet, Modbus")
        custom_rating = st.text_input("Screen Size (inch)", placeholder="e.g. 10.4")

    elif comp_type == "Encoder / Resolver":
        voltage     = st.text_input("Supply Voltage (V)", placeholder="e.g. 24 DC")
        ppr         = st.text_input("PPR / Bits / Turns", placeholder="e.g. 1024 PPR")
        output_type = st.text_input("Output Type", placeholder="e.g. HTL, TTL, SSI")

    elif comp_type == "Load Cell":
        voltage       = st.text_input("Supply Voltage (V)", placeholder="e.g. 10 DC")
        capacity      = st.text_input("Capacity (kg / N)", placeholder="e.g. 500 kg")
        custom_rating = st.text_input("Output (mV/V)", placeholder="e.g. 2 mV/V")

    elif comp_type == "Circuit Breaker / MCCB":
        amps         = st.text_input("Rated Current (A)", placeholder="e.g. 63")
        voltage      = st.text_input("Rated Voltage (V)", placeholder="e.g. 415")
        breaking_cap = st.text_input("Breaking Capacity (kA)", placeholder="e.g. 25")
        poles        = st.text_input("No. of Poles", placeholder="e.g. 3")

    elif comp_type == "Relay / Contactor":
        voltage = st.text_input("Coil Voltage (V)", placeholder="e.g. 220 AC")
        amps    = st.text_input("Contact Rating (A)", placeholder="e.g. 32")

    elif comp_type == "Transformer":
        kva    = st.text_input("Rating (kVA)", placeholder="e.g. 100")
        prim_v = st.text_input("Primary Voltage (V)", placeholder="e.g. 11000")
        sec_v  = st.text_input("Secondary Voltage (V)", placeholder="e.g. 400")
        freq   = st.text_input("Frequency (Hz)", placeholder="e.g. 50")

    else:  # Other
        custom_rating = st.text_input("Rating / Specs", placeholder="Any relevant rating")
        voltage       = st.text_input("Supply Voltage (V)", placeholder="e.g. 24 DC")

    notes = st.text_area("Notes / Remarks", placeholder="Condition, location detail, etc.", height=80)

    st.divider()
    add_clicked = st.button("＋  Add Component", use_container_width=True, type="primary")

    if add_clicked:
        if not department.strip() or not machine.strip():
            st.error("Department and Machine are required.")
        elif not comp_type:
            st.error("Select a component type.")
        else:
            entry = {
                "id":           len(st.session_state.parts),
                "department":   department.strip(),
                "machine":      machine.strip(),
                "type":         comp_type,
                "tag_panel":    tag_panel.strip(),
                "tag_schematic":tag_schematic.strip(),
                "make":         make.strip(),
                "model":        model.strip(),
                "kw":           kw.strip()           if kw else "",
                "amps":         amps.strip()          if amps else "",
                "voltage":      voltage.strip()       if voltage else "",
                "rpm":          rpm.strip()           if rpm else "",
                "poles":        poles.strip()         if poles else "",
                "freq":         freq.strip()          if freq else "",
                "ins_class":    ins_class.strip()     if ins_class else "",
                "ip_rating":    ip_rating.strip()     if ip_rating else "",
                "io_count":     io_count.strip()      if io_count else "",
                "comm":         comm.strip()          if comm else "",
                "ppr":          ppr.strip()           if ppr else "",
                "output_type":  output_type.strip()   if output_type else "",
                "capacity":     capacity.strip()      if capacity else "",
                "breaking_cap": breaking_cap.strip()  if breaking_cap else "",
                "kva":          kva.strip()           if kva else "",
                "prim_v":       prim_v.strip()        if prim_v else "",
                "sec_v":        sec_v.strip()         if sec_v else "",
                "custom_rating":custom_rating.strip() if custom_rating else "",
                "notes":        notes.strip(),
            }
            st.session_state.parts.append(entry)
            st.success(f"✓  {comp_type} added — {machine.strip()}")
            st.rerun()

# ── Main Area ─────────────────────────────────────────────────────────────────
parts = st.session_state.parts

# Header row
col_title, col_clear = st.columns([5, 1])
with col_title:
    st.markdown("## Critical Parts List")
    st.markdown("<div style='color:#8fa5bc;font-size:0.85rem;margin-top:-0.6rem;'>Automation Panel Asset Register — Sapphire Fibres Limited</div>", unsafe_allow_html=True)
with col_clear:
    if parts and st.button("🗑 Clear All", help="Remove all logged parts"):
        st.session_state.parts = []
        st.rerun()

st.divider()

# ── Summary metrics ───────────────────────────────────────────────────────────
if parts:
    depts    = len(set(p["department"] for p in parts))
    machines = len(set(p["machine"]    for p in parts))
    total    = len(parts)
    st.markdown(f"""
    <div class="metric-row">
      <div class="metric-box"><div class="num">{total}</div><div class="lbl">Total Components</div></div>
      <div class="metric-box"><div class="num">{machines}</div><div class="lbl">Machines</div></div>
      <div class="metric-box"><div class="num">{depts}</div><div class="lbl">Departments</div></div>
    </div>
    """, unsafe_allow_html=True)

# ── Filter bar ────────────────────────────────────────────────────────────────
if parts:
    f1, f2, f3 = st.columns(3)
    all_depts    = sorted(set(p["department"] for p in parts))
    all_machines = sorted(set(p["machine"]    for p in parts))
    all_types    = sorted(set(p["type"]       for p in parts))

    with f1:
        filter_dept = st.selectbox("Filter by Department", ["All"] + all_depts, key="f_dept")
    with f2:
        filter_mach = st.selectbox("Filter by Machine",    ["All"] + all_machines, key="f_mach")
    with f3:
        filter_type = st.selectbox("Filter by Type",       ["All"] + all_types, key="f_type")

    filtered = [
        p for p in parts
        if (filter_dept == "All" or p["department"] == filter_dept)
        and (filter_mach == "All" or p["machine"]    == filter_mach)
        and (filter_type == "All" or p["type"]       == filter_type)
    ]
else:
    filtered = []

st.markdown("")  # spacer

# ── Parts list ────────────────────────────────────────────────────────────────
if not filtered:
    st.markdown("""
    <div class="empty-state">
        <div class="icon">🔌</div>
        <p><strong>No components logged yet.</strong><br>Use the sidebar to add your first component.</p>
    </div>
    """, unsafe_allow_html=True)
else:
    # Group by Department → Machine
    grouped = {}
    for p in filtered:
        grouped.setdefault(p["department"], {}).setdefault(p["machine"], []).append(p)

    for dept, machines in sorted(grouped.items()):
        st.markdown(f"<div class='sec-header'>📁 {dept}</div>", unsafe_allow_html=True)

        for machine_name, comps in sorted(machines.items()):
            with st.expander(f"🔧 {machine_name}  ({len(comps)} component{'s' if len(comps)!=1 else ''})", expanded=True):
                for p in comps:
                    badge_cls = BADGE_CLASS.get(p["type"], "badge-other")

                    # Build specs string
                    specs = []
                    if p["kw"]:           specs.append(f"**{p['kw']} kW**")
                    if p["amps"]:         specs.append(f"{p['amps']} A")
                    if p["voltage"]:      specs.append(f"{p['voltage']} V")
                    if p["rpm"]:          specs.append(f"{p['rpm']} RPM")
                    if p["kva"]:          specs.append(f"{p['kva']} kVA")
                    if p["prim_v"]:       specs.append(f"{p['prim_v']}→{p['sec_v']} V")
                    if p["breaking_cap"]: specs.append(f"{p['breaking_cap']} kA")
                    if p["capacity"]:     specs.append(f"{p['capacity']}")
                    if p["ppr"]:          specs.append(f"{p['ppr']}")
                    if p["comm"]:         specs.append(f"{p['comm']}")
                    if p["io_count"]:     specs.append(f"I/O: {p['io_count']}")
                    if p["ip_rating"]:    specs.append(f"{p['ip_rating']}")
                    if p["ins_class"]:    specs.append(f"Class {p['ins_class']}")
                    if p["custom_rating"]:specs.append(f"{p['custom_rating']}")
                    specs_str = "  ·  ".join(specs) if specs else "—"

                    tag_html = ""
                    if p["tag_panel"]:
                        tag_html += f'<span class="tag-chip">Panel: {p["tag_panel"]}</span>'
                    if p["tag_schematic"]:
                        tag_html += f'<span class="tag-chip">SLD: {p["tag_schematic"]}</span>'

                    card_col, del_col = st.columns([11, 1])
                    with card_col:
                        st.markdown(f"""
                        <div class="part-card">
                            <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">
                                <span class="badge {badge_cls}">{p['type']}</span>
                                <span style="font-weight:600;font-size:0.95rem;">{p['make']} {p['model']}</span>
                                <span style="margin-left:auto;color:#8fa5bc;font-size:0.8rem;">{tag_html}</span>
                            </div>
                            <div style="font-size:0.85rem;color:#4a6a88;margin-bottom:{'6px' if p['notes'] else '0'};">{specs_str}</div>
                            {'<div style="font-size:0.78rem;color:#8fa5bc;font-style:italic;">'+p["notes"]+'</div>' if p["notes"] else ""}
                        </div>
                        """, unsafe_allow_html=True)
                    with del_col:
                        st.markdown('<div class="del-btn">', unsafe_allow_html=True)
                        if st.button("✕", key=f"del_{p['id']}", help="Remove"):
                            st.session_state.parts = [x for x in st.session_state.parts if x["id"] != p["id"]]
                            st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)

# ── Excel Export ──────────────────────────────────────────────────────────────
if parts:
    st.divider()
    st.markdown("### Export")

    def build_excel(data):
        wb = Workbook()

        # ── Sheet 1: Full Data ───────────────────────────────────────────────
        ws = wb.active
        ws.title = "Critical Parts List"

        headers = [
            "Department", "Machine", "Component Type",
            "Make / Brand", "Model No.",
            "Tag (Panel)", "Tag (Schematic/SLD)",
            "Power (kW/kVA)", "Current (A)", "Voltage (V)",
            "Speed (RPM)", "Poles", "Freq (Hz)",
            "Ins. Class", "IP Rating",
            "Breaking Cap. (kA)", "Capacity", "I/O Count",
            "Protocol", "PPR / Output", "Other Rating",
            "Notes",
        ]

        # Header style
        hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill  = PatternFill("solid", fgColor="185FA5")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin      = Side(style="thin", color="D0DCEA")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border

        ws.row_dimensions[1].height = 28

        # Data rows
        fill_even = PatternFill("solid", fgColor="F4F8FF")
        fill_odd  = PatternFill("solid", fgColor="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        data_align= Alignment(vertical="top", wrap_text=True)

        for row_idx, p in enumerate(data, 2):
            row_data = [
                p["department"], p["machine"], p["type"],
                p["make"], p["model"],
                p["tag_panel"], p["tag_schematic"],
                p["kw"] or p["kva"],
                p["amps"],
                p["voltage"] or p["prim_v"],
                p["rpm"], p["poles"], p["freq"],
                p["ins_class"], p["ip_rating"],
                p["breaking_cap"], p["capacity"],
                p["io_count"], p["comm"],
                p["ppr"] or p["output_type"],
                p["custom_rating"],
                p["notes"],
            ]
            ws.append(row_data)
            fill = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font      = data_font
                cell.fill      = fill
                cell.alignment = data_align
                cell.border    = border

        # Column widths
        col_widths = [18,20,22,18,20,14,18,12,12,12,10,8,8,10,10,12,12,12,14,14,14,30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

        # ── Sheet 2: Summary ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Critical Parts List — Summary"])
        ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="185FA5")
        ws2.append([f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}"])
        ws2["A2"].font = Font(name="Calibri", italic=True, size=10, color="6B8BA4")
        ws2.append([])

        ws2.append(["Department", "Machine", "Component Type", "Count"])
        for c in [ws2["A4"], ws2["B4"], ws2["C4"], ws2["D4"]]:
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="185FA5")
            c.alignment = Alignment(horizontal="center")

        from collections import Counter
        counter = Counter((p["department"], p["machine"], p["type"]) for p in data)
        for (dept, mach, ctype), cnt in sorted(counter.items()):
            ws2.append([dept, mach, ctype, cnt])

        for col in ["A","B","C","D"]:
            ws2.column_dimensions[col].width = 24

        # ── Sheet 3: By Department ───────────────────────────────────────────
        ws3 = wb.create_sheet("By Department")
        ws3.append(["Department", "Total Components"])
        ws3["A1"].font = Font(bold=True); ws3["B1"].font = Font(bold=True)
        dept_counts = Counter(p["department"] for p in data)
        for dept, cnt in sorted(dept_counts.items()):
            ws3.append([dept, cnt])
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 18

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    col_exp1, col_exp2 = st.columns([2, 1])
    with col_exp1:
        fname = f"Critical_Parts_List_{datetime.now().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="⬇  Download Excel File",
            data=build_excel(parts),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_exp2:
        st.markdown(f"<div style='font-size:0.78rem;color:#8fa5bc;padding-top:0.6rem;'>{len(parts)} component{'s' if len(parts)!=1 else ''} · {len(set(p['machine'] for p in parts))} machine{'s' if len(set(p['machine'] for p in parts))!=1 else ''}</div>", unsafe_allow_html=True)
